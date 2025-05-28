import webview
import openpyxl # Import openpyxl for creating Excel files
import os # For path manipulation
import platform # For detecting OS to open files
import subprocess # For opening files on non-Windows OS
from datetime import datetime # For handling dates
import json # For saving vendor details

# Define a Python API class to expose functions to JavaScript
class Api:
    def __init__(self):
        self.window = None

    def set_window(self, window):
        """Set the window reference when window is created"""
        self.window = window

    def get_window_scaling_factor(self):
        """Returns the scaling factor of the current Pywebview window."""
        try:
            if self.window:
                return self.window.scaling
            return 1.0
        except Exception as e:
            print(f"Error getting window scaling factor: {e}")
            return 1.0

    def move_window(self, dx, dy):
        """Moves the Pywebview window by dx and dy pixels."""
        try:
            if self.window:
                x, y = self.window.x, self.window.y
                self.window.move(x + dx, y + dy)
        except Exception as e:
            print(f"Error moving window: {e}")

    def create_excel_file(self, project_name):
        """
        Creates an empty XLSX file with the given project name inside a 'projects' folder.
        """
        try:
            # Ensure the project name is safe for a filename
            safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '.', '_')).strip()
            if not safe_project_name:
                safe_project_name = "untitled_project" # Fallback name

            # Define the projects directory path
            projects_dir = "projects"
            
            # Create the 'projects' directory if it doesn't exist
            os.makedirs(projects_dir, exist_ok=True)

            file_name = f"{safe_project_name}.xlsx"
            file_path = os.path.join(projects_dir, file_name) # Combine directory and file name
            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            
            # Get the active sheet (default one)
            sheet = workbook.active
            sheet.title = "Sheet1" # You can rename it if needed

            # Save the workbook in the specified directory
            workbook.save(file_path)
            print(f"Successfully created {file_path}")
            return {"success": True, "message": f"Project '{file_name}' created successfully in '{projects_dir}' folder!"}
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            return {"success": False, "message": f"Failed to create project: {e}"}

    def list_excel_files(self):
        """
        Lists all .xlsx files in the 'projects' directory.
        """
        projects_dir = "projects"
        files = []
        if os.path.exists(projects_dir) and os.path.isdir(projects_dir):
            for f_name in os.listdir(projects_dir):
                if f_name.endswith(".xlsx"):
                    files.append(f_name)
        return files

    def open_excel_file(self, file_name):
        """
        Opens the specified Excel file from the 'projects' directory.
        """
        projects_dir = "projects"
        file_path = os.path.join(projects_dir, file_name)
        
        if not os.path.exists(file_path):
            return {"success": False, "message": f"File '{file_name}' not found."}

        try:
            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin": # macOS
                subprocess.call(["open", file_path])
            else: # Linux and other Unix-like systems
                subprocess.call(["xdg-open", file_path])
            print(f"Successfully opened {file_path}")
            return {"success": True, "message": f"Opened '{file_name}'."}
        except Exception as e:
            print(f"Error opening file: {e}")
            return {"success": False, "message": f"Failed to open file: {e}"}

    def get_po_sheets(self, project_name):
        """
        Gets all sheets from a project's Excel file.
        """
        try:
            projects_dir = "projects"
            file_path = os.path.join(projects_dir, project_name)
            
            if not os.path.exists(file_path):
                return {"success": False, "message": f"Project '{project_name}' not found."}

            # Load workbook in read-only mode with data_only=True
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Get all sheet names
            sheets = workbook.sheetnames
            
            # Sort sheets numerically if they contain numbers
            def get_sheet_number(sheet_name):
                try:
                    # Extract number from sheet name (e.g., "123" -> 123)
                    num = ''.join(filter(str.isdigit, sheet_name))
                    return int(num) if num else float('inf')
                except ValueError:
                    return float('inf')
            
            sheets.sort(key=get_sheet_number)
            
            workbook.close()
            print(f"Found {len(sheets)} sheets in {project_name}")  # Debug print
            return {"success": True, "sheets": sheets}
        except Exception as e:
            print(f"Error getting sheets: {e}")
            return {"success": False, "message": str(e)}

    def create_po(self, project_name, po_name):
        """
        Creates a new PO in the project's Excel file.
        """
        try:
            projects_dir = "projects"
            file_path = os.path.join(projects_dir, project_name)
            
            if not os.path.exists(file_path):
                return {"success": False, "message": f"Project '{project_name}' not found."}

            # Load the template
            template_path = "Final PO Format.xlsx"
            if not os.path.exists(template_path):
                return {"success": False, "message": "PO template not found."}

            # Load both workbooks
            project_wb = openpyxl.load_workbook(file_path)
            template_wb = openpyxl.load_workbook(template_path)

            # Check if PO name already exists
            if po_name in project_wb.sheetnames:
                return {"success": False, "message": f"PO '{po_name}' already exists."}

            # Copy template sheet to project workbook
            template_sheet = template_wb.active
            new_sheet = project_wb.create_sheet(po_name)
            
            # Copy all cells from template
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    
                    # Copy styles safely by creating new style objects
                    if cell.has_style:
                        try:
                            # Create new style objects instead of copying references
                            new_cell.font = openpyxl.styles.Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                vertAlign=cell.font.vertAlign,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color=cell.font.color
                            )
                            
                            new_cell.border = openpyxl.styles.Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                            
                            new_cell.fill = openpyxl.styles.PatternFill(
                                fill_type=cell.fill.fill_type,
                                start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color
                            )
                            
                            new_cell.number_format = cell.number_format
                            
                            new_cell.alignment = openpyxl.styles.Alignment(
                                horizontal=cell.alignment.horizontal,
                                vertical=cell.alignment.vertical,
                                text_rotation=cell.alignment.text_rotation,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                        except Exception as style_error:
                            print(f"Warning: Could not copy some styles: {style_error}")
                            continue

            # Save the project workbook
            project_wb.save(file_path)
            return {"success": True, "message": f"PO '{po_name}' created successfully."}
        except Exception as e:
            print(f"Error creating PO: {e}")
            return {"success": False, "message": str(e)}

    def delete_po(self, project_name, po_name):
        """
        Marks a PO as deleted by coloring its tab red.
        """
        try:
            projects_dir = "projects"
            file_path = os.path.join(projects_dir, project_name)
            
            if not os.path.exists(file_path):
                return {"success": False, "message": f"Project '{project_name}' not found."}

            workbook = openpyxl.load_workbook(file_path)
            
            if po_name not in workbook.sheetnames:
                return {"success": False, "message": f"PO '{po_name}' not found."}

            # Mark sheet as deleted by coloring tab red
            sheet = workbook[po_name]
            sheet.sheet_properties.tabColor = "FFFF0000"  # Red color
            
            workbook.save(file_path)
            return {"success": True, "message": f"PO '{po_name}' marked as deleted."}
        except Exception as e:
            print(f"Error deleting PO: {e}")
            return {"success": False, "message": str(e)}

    def get_vendors(self):
        """
        Gets all unique vendors from all POs across all projects.
        """
        try:
            vendors = set()
            projects_dir = "projects"
            
            if not os.path.exists(projects_dir):
                return {"success": True, "vendors": []}

            for file_name in os.listdir(projects_dir):
                if file_name.endswith(".xlsx"):
                    file_path = os.path.join(projects_dir, file_name)
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    
                    for sheet in workbook.sheetnames:
                        if sheet.startswith("PO"):
                            sheet_obj = workbook[sheet]
                            # Assuming vendor name is in cell B2 (adjust as needed)
                            vendor = sheet_obj["B2"].value
                            if vendor:
                                vendors.add(vendor)
                    
                    workbook.close()
            
            return {"success": True, "vendors": sorted(list(vendors))}
        except Exception as e:
            print(f"Error getting vendors: {e}")
            return {"success": False, "message": str(e)}

    def get_vendor_details(self, vendor_name):
        """
        Gets the most recent details for a vendor from any PO.
        """
        try:
            vendor_details = {
                "name": vendor_name,
                "address": "",
                "contact": "",
                "email": ""
            }
            
            projects_dir = "projects"
            if not os.path.exists(projects_dir):
                return {"success": True, "details": vendor_details}

            # Find the most recent PO with this vendor
            latest_date = None
            for file_name in os.listdir(projects_dir):
                if file_name.endswith(".xlsx"):
                    file_path = os.path.join(projects_dir, file_name)
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    
                    for sheet in workbook.sheetnames:
                        if sheet.startswith("PO"):
                            sheet_obj = workbook[sheet]
                            if sheet_obj["B2"].value == vendor_name:
                                # Assuming PO date is in cell B3 (adjust as needed)
                                po_date = sheet_obj["B3"].value
                                if isinstance(po_date, datetime) and (latest_date is None or po_date > latest_date):
                                    latest_date = po_date
                                    # Get vendor details (adjust cell references as needed)
                                    vendor_details["address"] = sheet_obj["B4"].value or ""
                                    vendor_details["contact"] = sheet_obj["B5"].value or ""
                                    vendor_details["email"] = sheet_obj["B6"].value or ""
                    
                    workbook.close()
            
            return {"success": True, "details": vendor_details}
        except Exception as e:
            print(f"Error getting vendor details: {e}")
            return {"success": False, "message": str(e)}

    def get_items_for_vendor(self, vendor_name):
        """
        Gets all items previously ordered from a specific vendor.
        """
        try:
            items = set()
            projects_dir = "projects"
            
            if not os.path.exists(projects_dir):
                return {"success": True, "items": []}

            for file_name in os.listdir(projects_dir):
                if file_name.endswith(".xlsx"):
                    file_path = os.path.join(projects_dir, file_name)
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    
                    for sheet in workbook.sheetnames:
                        if sheet.startswith("PO"):
                            sheet_obj = workbook[sheet]
                            if sheet_obj["B2"].value == vendor_name:
                                # Assuming items start from row 10 (adjust as needed)
                                for row in range(10, sheet_obj.max_row + 1):
                                    item = sheet_obj[f"A{row}"].value
                                    if item:
                                        items.add(item)
                    
                    workbook.close()
            
            return {"success": True, "items": sorted(list(items))}
        except Exception as e:
            print(f"Error getting items: {e}")
            return {"success": False, "message": str(e)}

    def save_po_data(self, project_name, po_name, po_data):
        """
        Saves PO data to the specified sheet.
        """
        try:
            projects_dir = "projects"
            file_path = os.path.join(projects_dir, project_name)
            
            if not os.path.exists(file_path):
                return {"success": False, "message": f"Project '{project_name}' not found."}

            workbook = openpyxl.load_workbook(file_path)
            
            if po_name not in workbook.sheetnames:
                return {"success": False, "message": f"PO '{po_name}' not found."}

            sheet = workbook[po_name]
            
            # Save vendor details
            sheet["B2"] = po_data["vendor"]["name"]
            sheet["B4"] = po_data["vendor"]["address"]
            sheet["B5"] = po_data["vendor"]["contact"]
            sheet["B6"] = po_data["vendor"]["email"]
            
            # Save delivery details
            sheet["B3"] = po_data["delivery"]["date"]
            sheet["B7"] = po_data["delivery"]["instructions"]
            
            # Save items
            row = 10  # Starting row for items
            for item in po_data["items"]:
                sheet[f"A{row}"] = item["name"]
                sheet[f"B{row}"] = item["quantity"]
                sheet[f"C{row}"] = item["unit_price"]
                sheet[f"D{row}"] = item["description"]
                row += 1
            
            # Save terms and conditions
            sheet["B8"] = po_data["terms"]
            
            workbook.save(file_path)
            return {"success": True, "message": "PO data saved successfully."}
        except Exception as e:
            print(f"Error saving PO data: {e}")
            return {"success": False, "message": str(e)}

    def save_vendor_details(self, vendor_name, vendor_details):
        """
        Saves vendor details to a JSON file for future use.
        """
        try:
            vendors_dir = "vendors"
            os.makedirs(vendors_dir, exist_ok=True)
            
            vendor_file = os.path.join(vendors_dir, "vendor_details.json")
            
            # Load existing vendors if file exists
            existing_vendors = {}
            if os.path.exists(vendor_file):
                with open(vendor_file, 'r') as f:
                    existing_vendors = json.load(f)
            
            # Update or add new vendor
            existing_vendors[vendor_name] = vendor_details
            
            # Save updated vendors
            with open(vendor_file, 'w') as f:
                json.dump(existing_vendors, f, indent=4)
            
            return {"success": True, "message": f"Vendor '{vendor_name}' details saved successfully."}
        except Exception as e:
            print(f"Error saving vendor details: {e}")
            return {"success": False, "message": str(e)}

    def minimize_window(self):
        """Minimizes the window."""
        webview.active_window().minimize()

    def toggle_maximize(self):
        """Toggles between maximized and normal state."""
        window = webview.active_window()
        if window.fullscreen:
            window.toggle_fullscreen()
        else:
            window.toggle_fullscreen()

    def close_window(self):
        """Closes the window."""
        webview.active_window().destroy()

# Instantiate the API class
api = Api()

try:
    with open("inter_base64.txt", "r") as f:
        inter_font_base64 = f.read().strip()
except FileNotFoundError:
    print("Warning: 'inter_base64.txt' not found. The custom Inter font will not load.")
    inter_font_base64 = "" 

html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>Purchase</title>
    <style>
        @font-face {{
            font-family: 'Inter';
            src: url(data:font/truetype;charset=utf-8;base64,{inter_font_base64}) format('truetype');
            font-weight: normal;
            font-style: normal;
        }}
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        html, body {{
            width: 100%;
            height: 100%;
            background: linear-gradient(270deg, #014574, #5228AC, #221439);
            background-size: 400% 400%;
            animation: gradientAnimation 90s ease infinite;
            font-family: 'Inter', Arial, sans-serif;
            overflow: hidden;
        }}

        @keyframes gradientAnimation {{
            0% {{ background-position: 0% 50%; }}
            50% {{ background-position: 100% 50%; }}
            100% {{ background-position: 0% 50%; }}
        }}

        /* Custom Title Bar */
        .custom-title-bar {{
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            height: 32px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-bottom: 1px solid rgba(255, 255, 255, 0.15);
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 0 10px;
            z-index: 1000;
            -webkit-app-region: drag;
        }}

        .title-bar-title {{
            color: #E3EEE9;
            font-family: 'Inter', Arial, sans-serif;
            font-size: 14px;
            margin-left: 10px;
            -webkit-app-region: drag;
        }}

        .title-bar-controls {{
            display: flex;
            gap: 8px;
            -webkit-app-region: no-drag;
        }}

        .title-bar-button {{
            width: 24px;
            height: 24px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            display: flex;
            align-items: center;
            justify-content: center;
            cursor: pointer;
            transition: all 0.2s ease;
            color: #E3EEE9;
            font-size: 16px;
            line-height: 1;
            -webkit-app-region: no-drag;
        }}

        .title-bar-button:hover {{
            background: rgba(255, 255, 255, 0.2);
            transform: translateY(-1px);
        }}

        .title-bar-button:active {{
            transform: translateY(1px);
        }}

        .close-button:hover {{
            background: rgba(255, 0, 0, 0.2);
        }}

        .minimize-button:hover {{
            background: rgba(255, 255, 255, 0.2);
        }}

        .maximize-button:hover {{
            background: rgba(255, 255, 255, 0.2);
        }}

        /* Add padding to container to account for title bar */
        .container {{
            margin-top: 32px; /* Match title bar height */
            display: flex;
            justify-content: center;
            align-items: center;
            height: calc(100% - 32px); /* Subtract title bar height */
            padding: 30px;
            gap: 30px;
            position: relative;
            z-index: 1;
            transition: transform 0.3s ease-out, opacity 0.2s ease-out;
        }}

        .glass-button {{
            flex: 1;
            height: 100%;
            background: #9859FF;
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 24px;
            font-family: 'Inter', monospace;
            font-size: 28px;
            font-weight: normal;
            cursor: pointer;
            transition: box-shadow 0.075s, transform 0.075s;
            position: relative;
            z-index: 2;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 0 16px;
            user-select: none;
            white-space: nowrap;
            will-change: box-shadow, transform, border-color, opacity;
            text-align: center;
            text-decoration: none;
            background-image: linear-gradient(135deg, #93A5CF, #E3EEE9);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        .glass-button:focus {{
            outline: none;
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 0 0 1.5px inset,
                rgba(45, 35, 66, 0.15) 0 2px 4px,
                rgba(45, 35, 66, 0.1) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.1) 0 -3px 0 inset;
        }}

        .glass-button:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px;
            transform: translateY(-2px);
        }}

        .glass-button:active {{
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
            transform: translateY(2px);
        }}

        .cursor-bubble {{
            position: absolute;
            width: 600px;
            height: 600px;
            border-radius: 70%;
            background: #69E6FE;
            opacity: 0.6;
            box-shadow:
                0 0 150px rgba(105, 230, 254, 0.9),
                0 0 300px rgba(105, 230, 254, 0.7),
                0 0 450px rgba(105, 230, 254, 0.5);
            pointer-events: none;
            animation: gentleWobble 5s ease-in-out infinite;
            z-index: 0;
        }}

        @keyframes gentleWobble {{
            0%, 100% {{ border-radius: 50% 50% 50% 50%; }}
            25% {{ border-radius: 50% 45% 50% 55%; }}
            50% {{ border-radius: 50% 50% 45% 55%; }}
            75% {{ border-radius: 55% 45% 50% 50%; }}
        }}

        /* Adjust overlay positions to account for title bar */
        .overlay-common-style {{
            position: fixed;
            top: 42px; /* Title bar height (32px) + 10px spacing */
            left: 8px;
            right: 8px;
            bottom: 8px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 25px;
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            opacity: 0;
            pointer-events: none;
            z-index: 10;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }}

        #overlayRect {{
            transform: scale(0.1);
            transition: opacity 0.25s ease-out, transform 0.25s cubic-bezier(0.25, 0.46, 0.45, 0.94);
        }}

        #overlayRect.overlay-visible {{
            opacity: 1;
            transform: scale(1);
            pointer-events: auto;
        }}

        #existingProjectOverlay {{
            transform: scale(1.8);
            opacity: 0;
            transition: opacity 0.2s ease-out, transform 0.25s cubic-bezier(0.25, 0.46, 0.45, 0.94);
        }}

        #existingProjectOverlay.overlay-visible {{
            opacity: 1;
            transform: scale(1);
            pointer-events: auto;
        }}

        .glass-search-input {{
            width: 80%;
            max-width: 500px;
            height: 50px;
            margin-top: 20px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 24px;
            font-family: 'Inter', monospace;
            font-size: 18px;
            font-weight: normal;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            padding: 0 20px;
            outline: none;
            transition: box-shadow 0.15s ease, transform 0.15s ease;
            text-align: left;
            color: #E3EEE9;
        }}

        .glass-search-input::placeholder {{
            color: rgba(54, 57, 90, 0.6);
        }}

        .glass-search-input:focus {{
            outline: none;
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 0 0 1.5px inset,
                rgba(45, 35, 66, 0.15) 0 2px 4px,
                rgba(45, 35, 66, 0.1) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.1) 0 -3px 0 inset;
        }}

        .glass-search-input:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.12) 0 -3px 0 inset;
            transform: translateY(-2px);
        }}

        .glass-search-input:active {{
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
            transform: translateY(2px);
        }}

        /* Adjust back button position */
        .back-button-common-style {{
            position: fixed;
            top: 44px; /* Title bar height + 12px */
            left: 12px;
            width: 40px;
            height: 40px;
            border-radius: 25px;
            background: #9859FF;
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            font-size: 24px;
            line-height: 36px;
            text-align: center;
            cursor: pointer;
            opacity: 0;
            pointer-events: none;
            user-select: none;
            transition: opacity 0.2s ease, box-shadow 0.075s ease, transform 0.075s ease;
            z-index: 20;
            background-image: linear-gradient(135deg, #93A5CF, #E3EEE9);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        .back-button-common-style.visible {{
            opacity: 1;
            pointer-events: auto;
        }}

        .back-button-common-style:focus {{
            outline: none;
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 0 0 1.5px inset,
                rgba(45, 35, 66, 0.15) 0 2px 4px,
                rgba(45, 35, 66, 0.1) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.1) 0 -3px 0 inset;
        }}

        .back-button-common-style.visible:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.12) 0 -3px 0 inset;
            transform: translateY(-2px);
        }}

        .back-button-common-style.visible:active {{
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
            transform: translateY(2px);
        }}

        .slide-button {{
            outline: 0;
            border: 0;
            display: flex;
            flex-direction: column;
            width: 100%;
            max-width: 200px;
            height: 50px;
            border-radius: 24px;
            box-shadow: 0 0.625em 1em 0 rgba(30, 143, 255, 0.35);
            overflow: hidden;
            margin-top: 20px;
            cursor: pointer;
            opacity: 0;
            pointer-events: none;
            transform: translateY(10px);
            transition: opacity 0.3s ease-out, transform 0.3s ease-out, box-shadow 0.3s ease;
        }}

        .slide-button.visible {{
            opacity: 1;
            pointer-events: auto;
            transform: translateY(0px);
        }}

        .slide-button:hover {{
            box-shadow: 0 0.8em 1.2em 0 rgba(30, 143, 255, 0.45);
            transform: translateY(-2px);
        }}

        .slide-button.success-state {{
            background-color: #21dc62;
            box-shadow: 0 0.625em 1em 0 rgba(33, 220, 98, 0.35);
            transform: translateY(2px);
            transition: background-color 0.3s ease, box-shadow 0.3s ease, transform 0.1s ease;
        }}

        .slide-button.success-state div {{
            transform: translateY(-50px);
        }}

        .slide-button div {{
            transform: translateY(0px);
            width: 100%;
            transition: 0.6s cubic-bezier(.16,1,.3,1);
        }}

        .slide-button div span {{
            display: flex;
            align-items: center;
            justify-content: center;
            height: 50px;
            padding: 0.75em 1.125em;
        }}

        .slide-button div:nth-child(1) {{
            background-color: #1e90ff; /* Original blue background */
        }}

        .slide-button div:nth-child(2) {{
            background-color: #21dc62; /* Green background for the second div */
        }}

        .slide-button p {{
            font-size: 17px;
            font-weight: bold;
            color: #ffffff;
            font-family: 'Inter', Arial, sans-serif; /* Ensure Inter font */
            text-align: center; /* Center align the text */
        }}

        .slide-button:active {{
            transform: scale(0.95);
        }}

        /* Styles for the project list container */
        #projectListContainer {{
            width: 80%;
            max-width: 500px;
            height: 70%; /* Take up most of the height */
            margin-top: 20px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 25px;
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            overflow-y: auto; /* Enable scrolling for long lists */
            scrollbar-width: none; /* For Firefox */
            -ms-overflow-style: none; /* For Internet Explorer and Edge */
            padding: 10px;
            display: flex;
            flex-direction: column;
            gap: 10px; /* Space between list items */
        }}
        /* For Webkit browsers (Chrome, Safari) */
        #projectListContainer::-webkit-scrollbar {{
            display: none;
        }}

        /* Style for individual project items */
        .project-item {{
            background: #9859FF; /* Same background as main buttons */
            border-radius: 15px;
            padding: 10px 15px;
            cursor: pointer;
            font-family: 'Inter', Arial, sans-serif;
            font-size: 16px;
            font-weight: bold;
            text-align: center;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            /* Added transitions for animation */
            transition: opacity 0.3s ease-out, transform 0.3s ease-out, height 0.3s ease-out, margin 0.3s ease-out, padding 0.3s ease-out, box-shadow 0.15s, border-radius 0.15s;
            will-change: transform, opacity; /* Optimize for animation */

            /* Text gradient for project items */
            background-image: linear-gradient(135deg, #2B2B2B, #575757); /* New gradient colors */
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        .project-item:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.12) 0 -3px 0 inset;
            transform: translateY(-2px);
        }}

        .project-item:active {{
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
            transform: translateY(2px);
        }}

        .project-item.hidden {{ /* For items that are filtered out */
            opacity: 0;
            height: 0 !important; /* Force height to 0 */
            margin-top: 0 !important;
            margin-bottom: 0 !important;
            padding-top: 0 !important;
            padding-bottom: 0 !important;
            pointer-events: none;
            /* Removed transform: translateY(10px); for quick fade */
        }}

        #noProjectsMessage {{
            color: #E3EEE9;
            font-family: 'Inter', Arial, sans-serif;
            font-size: 18px;
            text-align: center;
            margin-top: 20px;
        }}

        /* Adjust PO Management overlay position */
        .po-management-overlay {{
            position: fixed;
            top: 42px; /* Title bar height (32px) + 10px spacing */
            left: 8px;
            right: 8px;
            bottom: 8px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 25px;
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            opacity: 0;
            pointer-events: none;
            z-index: 10;
            display: flex;
            flex-direction: column;
            align-items: center;
            padding: 20px;
            transform: scale(1.8);
            transition: opacity 0.2s ease-out, transform 0.25s cubic-bezier(0.25, 0.46, 0.45, 0.94);
        }}

        .po-management-overlay.overlay-visible {{
            opacity: 1;
            transform: scale(1);
            pointer-events: auto;
        }}

        .po-list-container {{
            width: 80%;
            max-width: 500px;
            height: calc(70% - 10px); /* Adjust height to account for spacing */
            margin-top: 20px;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 25px;
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            overflow-y: auto;
            padding: 10px;
            display: flex;
            flex-direction: column;
            gap: 10px;
            scrollbar-width: none; /* For Firefox */
            -ms-overflow-style: none; /* For Internet Explorer and Edge */
        }}

        /* For Webkit browsers (Chrome, Safari) */
        .po-list-container::-webkit-scrollbar {{
            display: none;
        }}

        .po-item {{
            background: #9859FF;
            border-radius: 15px;
            padding: 10px 15px;
            cursor: pointer;
            font-family: 'Inter', Arial, sans-serif;
            font-size: 16px;
            font-weight: bold;
            text-align: center;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            transition: all 0.3s ease;
            background-image: linear-gradient(135deg, #2B2B2B, #575757);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        .po-item:hover {{
            transform: translateY(-2px);
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px;
        }}

        .po-item:active {{
            transform: translateY(2px);
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
        }}

        .add-po-button {{
            position: fixed;
            bottom: 20px;
            right: 20px;
            width: 50px;
            height: 50px;
            border-radius: 25px;
            background: #9859FF;
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 2px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px;
            font-size: 24px;
            cursor: pointer;
            transition: box-shadow 0.075s, transform 0.075s;
            display: flex;
            align-items: center;
            justify-content: center;
            user-select: none;
            will-change: box-shadow, transform, border-color, opacity;
            background-image: linear-gradient(135deg, #93A5CF, #E3EEE9);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        .add-po-button:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px;
            transform: translateY(-2px);
        }}

        .add-po-button:active {{
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 3px 7px inset;
            transform: translateY(2px);
        }}

        /* Excel file opening button */
        .open-excel-button {{
            width: fit-content;
            min-width: 200px;
            height: 50px;
            margin: 20px auto;
            background: #9859FF;
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 24px;
            font-family: 'Inter', monospace;
            font-size: 18px;
            font-weight: normal;
            cursor: pointer;
            transition: box-shadow 0.075s, transform 0.075s;
            position: relative;
            z-index: 2;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 0 30px;
            user-select: none;
            white-space: nowrap;
            will-change: box-shadow, transform, border-color, opacity;
            text-align: center;
            text-decoration: none;
            background-image: linear-gradient(135deg, #93A5CF, #E3EEE9);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
        }}

        /* PO Search input */
        .po-search-input {{
            width: 80%;
            max-width: 500px;
            height: 50px;
            margin: 0 auto 20px auto;
            background: rgba(252, 252, 253, 0.15);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border-radius: 24px;
            font-family: 'Inter', monospace;
            font-size: 16px;
            font-weight: normal;
            border: 1px solid rgba(255, 255, 255, 0.15);
            box-shadow:
                rgba(45, 35, 66, 0.1) 0 2px 4px,
                rgba(45, 35, 66, 0.08) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.06) 0 -3px 0 inset;
            padding: 0 20px;
            outline: none;
            transition: box-shadow 0.15s ease, transform 0.15s ease;
            text-align: left;
            color: #E3EEE9;
        }}

        .po-search-input::placeholder {{
            color: rgba(227, 238, 233, 0.6);
        }}

        .po-search-input:focus {{
            outline: none;
            box-shadow:
                rgba(214, 214, 231, 0.3) 0 0 0 1.5px inset,
                rgba(45, 35, 66, 0.15) 0 2px 4px,
                rgba(45, 35, 66, 0.1) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.1) 0 -3px 0 inset;
        }}

        .po-search-input:hover {{
            box-shadow:
                rgba(45, 35, 66, 0.2) 0 4px 8px,
                rgba(45, 35, 66, 0.15) 0 7px 13px -3px,
                rgba(214, 214, 231, 0.12) 0 -3px 0 inset;
            transform: translateY(-2px);
        }}
    </style>
</head>
<body>
    <div class="custom-title-bar pywebview-drag-region">
        <div class="title-bar-title">Purchase</div>
        <div class="title-bar-controls">
            <button class="title-bar-button minimize-button" id="minimizeButton">-</button>
            <button class="title-bar-button maximize-button" id="maximizeButton">□</button>
            <button class="title-bar-button close-button" id="closeButton">×</button>
        </div>
    </div>
    <div class="cursor-bubble" id="cursorBubble"></div>
    <div class="container" id="mainContainer">
        <button id="existingBtn" class="glass-button">Existing Project</button>
        <button id="newBtn" class="glass-button">New Project</button>
    </div>

    <div id="overlayRect" class="overlay-common-style">
        <button id="backBtn" class="back-button-common-style">&#8592;</button>
        <input type="text" id="projectSearch" class="glass-search-input" placeholder="Enter Project Name">
        <button id="createProjectBtn" class="slide-button">
            <div>
                <span><p id="createProjectText">Create Project</p></span>
            </div>
            <div>
                <span><p>Project made</p></span> </div>
        </button>
    </div>

    <div id="existingProjectOverlay" class="overlay-common-style">
        <button id="existingBackBtn" class="back-button-common-style">&#8592;</button>
        <input type="text" id="existingProjectSearch" class="glass-search-input" placeholder="Search Existing Projects">
        <p id="noProjectsMessage" style="display: none;">No existing projects found.</p>
        <div id="projectListContainer">
        </div>
    </div>

    <!-- PO Management Window -->
    <div id="poManagementOverlay" class="po-management-overlay">
        <button id="poBackBtn" class="back-button-common-style">&#8592;</button>
        <h2 id="poProjectName" style="color: #E3EEE9; margin-bottom: 20px;"></h2>
        <input type="text" id="poSearch" class="po-search-input" placeholder="Search POs...">
        <div id="poListContainer" class="po-list-container"></div>
        <button id="addPoButton" class="add-po-button">+</button>
        <button id="openExcelButton" class="open-excel-button">Open Excel File</button>
    </div>

    <!-- PO Data Entry Window -->
    <div id="poDataEntryOverlay" class="po-data-entry-overlay">
        <button id="poDataBackBtn" class="back-button-common-style">&#8592;</button>
        <h2 id="poDataTitle" style="color: #E3EEE9; margin-bottom: 20px;"></h2>
        
        <!-- Step 1: Vendor and Delivery Details -->
        <div id="step1" class="step-container">
            <div class="form-group">
                <label for="vendor">Vendor:</label>
                <select name="vendor" class="form-control">
                    <option value="">Select a vendor</option>
                    <option value="new">+ Add New Vendor</option>
                </select>
            </div>
            <div class="new-vendor-fields" style="display: none;">
                <div class="form-group new-vendor-field" style="display: none;">
                    <label for="newVendorName">Vendor Name:</label>
                    <input type="text" name="newVendorName" class="form-control">
                </div>
                <div class="form-group new-vendor-field" style="display: none;">
                    <label for="newVendorAddress">Vendor Address:</label>
                    <input type="text" name="newVendorAddress" class="form-control">
                </div>
                <div class="form-group new-vendor-field" style="display: none;">
                    <label for="newVendorContact">Contact Number:</label>
                    <input type="text" name="newVendorContact" class="form-control">
                </div>
                <div class="form-group new-vendor-field" style="display: none;">
                    <label for="newVendorEmail">Email:</label>
                    <input type="email" name="newVendorEmail" class="form-control">
                </div>
            </div>
            <div class="vendor-details">
                <div class="form-group">
                    <label for="vendorAddress">Vendor Address:</label>
                    <input type="text" name="vendorAddress" class="form-control" readonly>
                </div>
                <div class="form-group">
                    <label for="vendorContact">Contact Number:</label>
                    <input type="text" name="vendorContact" class="form-control" readonly>
                </div>
                <div class="form-group">
                    <label for="vendorEmail">Email:</label>
                    <input type="email" name="vendorEmail" class="form-control" readonly>
                </div>
            </div>
            <div class="form-group">
                <label for="deliveryDate">Delivery Date</label>
                <input type="date" id="deliveryDate" class="form-control">
            </div>
            <div class="form-group">
                <label for="deliveryInstructions">Delivery Instructions</label>
                <textarea id="deliveryInstructions" class="form-control" rows="3"></textarea>
            </div>
        </div>

        <!-- Step 2: Order Details -->
        <div id="step2" class="step-container" style="display: none;">
            <div id="itemsContainer">
                <div class="item-row">
                    <select class="form-control item-select">
                        <option value="">Select an item</option>
                    </select>
                    <input type="number" class="form-control item-quantity" placeholder="Qty">
                    <input type="number" class="form-control item-price" placeholder="Price">
                    <input type="text" class="form-control item-description" placeholder="Description">
                    <button class="remove-item-button">×</button>
                </div>
            </div>
            <button id="addItemButton" class="add-item-button">Add Item</button>
        </div>

        <!-- Step 3: Terms and Conditions -->
        <div id="step3" class="step-container" style="display: none;">
            <div class="form-group">
                <label>
                    <input type="radio" name="termsOption" value="default" checked>
                    Use Default Terms
                </label>
            </div>
            <div class="form-group">
                <label>
                    <input type="radio" name="termsOption" value="custom">
                    Enter Custom Terms
                </label>
            </div>
            <div class="form-group">
                <textarea id="termsText" class="form-control" rows="5" readonly></textarea>
            </div>
        </div>

        <!-- Navigation Buttons -->
        <div class="navigation-buttons">
            <button id="prevStepButton" class="nav-button" style="display: none;">Previous</button>
            <button id="nextStepButton" class="nav-button">Next</button>
        </div>
    </div>

    <script>
        // Global variables
        let allProjects = [];
        let currentProjectName = '';
        let allPOsInCurrentProject = [];
        let currentPoName = '';
        let globalVendorData = {{}};
        let globalItemData = [];
        let dpiScalingFactor = 1.0;  // Will be set when window loads
        let isDragging = false;
        let lastScreenX = 0;
        let lastScreenY = 0;

        document.addEventListener('DOMContentLoaded', function() {{
            // Initialize DPI scaling factor
            async function fetchScalingFactor() {{
                try {{
                    const result = await window.pywebview.api.get_window_scaling_factor();
                    if (typeof result === 'number') {{
                        dpiScalingFactor = result;
                        console.log('Window DPI Scaling Factor:', dpiScalingFactor);
                    }}
                }} catch (error) {{
                    console.error('Failed to fetch DPI scaling factor:', error);
                }}
            }}
            fetchScalingFactor();

            // Custom title bar dragging
            const titleBar = document.querySelector('.custom-title-bar');
            
            titleBar.addEventListener('mousedown', (e) => {{
                // Only drag with left mouse button and if not clicking on control buttons
                if (e.button === 0 && !e.target.closest('.title-bar-controls')) {{
                    isDragging = true;
                    lastScreenX = e.screenX;
                    lastScreenY = e.screenY;
                    e.preventDefault(); // Prevent text selection
                }}
            }});

            document.addEventListener('mousemove', (e) => {{
                if (!isDragging) return;
                
                const rawDx = e.screenX - lastScreenX;
                const rawDy = e.screenY - lastScreenY;
                
                // Apply DPI scaling factor
                const adjustedDx = rawDx / dpiScalingFactor;
                const adjustedDy = rawDy / dpiScalingFactor;
                
                window.pywebview.api.move_window(adjustedDx, adjustedDy);
                
                lastScreenX = e.screenX;
                lastScreenY = e.screenY;
            }});

            document.addEventListener('mouseup', () => {{
                isDragging = false;
            }});

            // Initialize all DOM elements
            const cursorBubble = document.getElementById('cursorBubble');
            const mainContainer = document.getElementById('mainContainer');
            const existingBtn = document.getElementById('existingBtn');
            const newBtn = document.getElementById('newBtn');
            const overlayRect = document.getElementById('overlayRect');
            const backBtn = document.getElementById('backBtn');
            const existingProjectOverlay = document.getElementById('existingProjectOverlay');
            const existingBackBtn = document.getElementById('existingBackBtn');
            const projectListContainer = document.getElementById('projectListContainer');
            const noProjectsMessage = document.getElementById('noProjectsMessage');
            const poManagementOverlay = document.getElementById('poManagementOverlay');
            const poBackBtn = document.getElementById('poBackBtn');
            const poProjectName = document.getElementById('poProjectName');
            const poListContainer = document.getElementById('poListContainer');
            const addPoButton = document.getElementById('addPoButton');

            // Initialize cursor bubble animation
            let mouseX = 0, mouseY = 0, bubbleX = 0, bubbleY = 0;

            document.addEventListener('mousemove', (e) => {{
                mouseX = e.clientX - cursorBubble.offsetWidth / 2;
                mouseY = e.clientY - cursorBubble.offsetHeight / 2;
            }});

            function updateBubblePosition() {{
                const dx = mouseX - bubbleX;
                const dy = mouseY - bubbleY;
                bubbleX += dx * 0.1;
                bubbleY += dy * 0.1;
                cursorBubble.style.transform = `translate(${bubbleX}px, ${bubbleY}px)`;
                requestAnimationFrame(updateBubblePosition);
            }}
            updateBubblePosition();

            // New Project button click handler
            newBtn.addEventListener('click', function() {{
                mainContainer.style.transition = 'transform 0.3s ease-out, opacity 0.2s ease-out';
                mainContainer.style.transform = 'scale(1.8)';
                mainContainer.style.opacity = '0';
                mainContainer.style.pointerEvents = 'none';

                setTimeout(() => {{
                    overlayRect.classList.add("overlay-visible");
                    backBtn.classList.add("visible");
                    document.getElementById('projectSearch').value = '';
                }}, 150);
            }});

            // Existing Project button click handler
            existingBtn.addEventListener('click', function() {{
                mainContainer.style.transition = 'transform 0.3s ease-out, opacity 0.2s ease-out';
                mainContainer.style.transform = 'scale(0.1)';
                mainContainer.style.opacity = '0';
                mainContainer.style.pointerEvents = 'none';

                setTimeout(() => {{
                    existingProjectOverlay.classList.add("overlay-visible");
                    existingBackBtn.classList.add("visible");
                    loadExistingProjects();
                }}, 150);
            }});

            // Back button handlers
            backBtn.addEventListener('click', function() {{
                overlayRect.classList.remove("overlay-visible");
                backBtn.classList.remove("visible");
                document.getElementById('projectSearch').value = '';

                setTimeout(() => {{
                    mainContainer.style.transition = 'transform 0.25s ease-in, opacity 0.25s ease-in';
                    mainContainer.style.transform = 'scale(1)';
                    mainContainer.style.opacity = '1';
                    mainContainer.style.pointerEvents = 'auto';
                }}, 25);
            }});

            existingBackBtn.addEventListener('click', function() {{
                existingProjectOverlay.classList.remove("overlay-visible");
                existingBackBtn.classList.remove("visible");
                projectListContainer.innerHTML = '';
                noProjectsMessage.style.display = 'none';
                document.getElementById('existingProjectSearch').value = '';

                setTimeout(() => {{
                    mainContainer.style.transition = 'transform 0.25s ease-in, opacity 0.25s ease-in';
                    mainContainer.style.transform = 'scale(1)';
                    mainContainer.style.opacity = '1';
                    mainContainer.style.pointerEvents = 'auto';
                }}, 25);
            }});

            // Project search input handler
            document.getElementById('projectSearch').addEventListener('input', function() {{
                const createProjectBtn = document.getElementById('createProjectBtn');
                if (this.value.trim() !== '') {{
                    createProjectBtn.classList.add('visible');
                }} else {{
                    createProjectBtn.classList.remove('visible');
                }}
            }});

            // Create project button handler
            document.getElementById('createProjectBtn').addEventListener('click', async function() {{
                const projectName = document.getElementById('projectSearch').value.trim();
                if (projectName) {{
                    this.classList.add('success-state');
                    this.querySelector('div:nth-child(1)').style.transform = 'translateY(-50px)';
                    this.style.pointerEvents = 'none';

                    const result = await window.pywebview.api.create_excel_file(projectName);
                    
                    if (result.success) {{
                        console.log(result.message);
                        setTimeout(() => {{
                            this.classList.remove('success-state');
                            this.querySelector('div:nth-child(1)').style.transform = 'translateY(0px)';
                            document.getElementById('projectSearch').value = '';
                            this.classList.remove('visible');
                            this.style.pointerEvents = 'auto';
                        }}, 800);
                    }} else {{
                        console.error("Error:", result.message);
                        this.classList.remove('success-state');
                        this.querySelector('div:nth-child(1)').style.transform = 'translateY(0px)';
                        this.style.pointerEvents = 'auto';
                    }}
                }}
            }});

            // Existing project search handler
            document.getElementById('existingProjectSearch').addEventListener('input', function() {{
                const searchTerm = this.value.toLowerCase().trim();
                const filteredProjects = allProjects.filter(name => 
                    name.toLowerCase().includes(searchTerm)
                );
                renderProjectList(filteredProjects);
            }});

            // Load existing projects
            async function loadExistingProjects() {{
                document.getElementById('existingProjectSearch').value = '';
                allProjects = await window.pywebview.api.list_excel_files();
                renderProjectList(allProjects);
            }}

            // Function to render project list
            function renderProjectList(projectsToDisplay) {{
                projectListContainer.innerHTML = '';
                projectsToDisplay.forEach(file_name => {{
                    const projectItem = document.createElement('div');
                    projectItem.classList.add('project-item');
                    // Remove .xlsx extension for display
                    projectItem.textContent = file_name.replace('.xlsx', '');
                    projectItem.addEventListener('click', () => {{
                        // Hide existing project overlay
                        existingProjectOverlay.classList.remove("overlay-visible");
                        existingBackBtn.classList.remove("visible");
                        
                        // Show PO Management Window
                        poProjectName.textContent = file_name.replace('.xlsx', '');
                        poManagementOverlay.classList.add("overlay-visible");
                        poBackBtn.classList.add("visible");
                        
                        // Load POs
                        window.pywebview.api.get_po_sheets(file_name).then(result => {{
                            if (result.success) {{
                                renderPOList(result.sheets);
                            }} else {{
                                console.error("Failed to load POs:", result.message);
                                poListContainer.innerHTML = '<div class="po-item">No POs found</div>';
                            }}
                        }});
                    }});
                    projectListContainer.appendChild(projectItem);
                }});

                // Update noProjectsMessage visibility
                noProjectsMessage.style.display = projectsToDisplay.length === 0 ? 'block' : 'none';
            }}

            // Function to render PO list
            function renderPOList(poSheets) {{
                poListContainer.innerHTML = '';
                console.log('Rendering PO list:', poSheets);  // Debug log
                
                if (poSheets && poSheets.length > 0) {{
                    poSheets.forEach(poName => {{
                        const poItem = document.createElement('div');
                        poItem.className = 'po-item';
                        poItem.textContent = poName;
                        poItem.addEventListener('click', () => {{
                            console.log('Clicked PO:', poName);
                        }});
                        poListContainer.appendChild(poItem);
                    }});
                }} else {{
                    const noPOsMessage = document.createElement('div');
                    noPOsMessage.className = 'po-item';
                    noPOsMessage.textContent = 'No POs found';
                    poListContainer.appendChild(noPOsMessage);
                }}
            }}

            // PO search input handler
            document.getElementById('poSearch').addEventListener('input', function() {{
                const searchTerm = this.value.toLowerCase().trim();
                const poItems = poListContainer.getElementsByClassName('po-item');
                
                Array.from(poItems).forEach(item => {{
                    const poName = item.textContent.toLowerCase();
                    if (poName.includes(searchTerm)) {{
                        item.style.display = '';
                    }} else {{
                        item.style.display = 'none';
                    }}
                }});
            }});

            // Event listener for PO Management back button
            poBackBtn.addEventListener('click', () => {{
                poManagementOverlay.classList.remove("overlay-visible");
                poBackBtn.classList.remove("visible");
                poListContainer.innerHTML = '';
                
                setTimeout(() => {{
                    existingProjectOverlay.classList.add("overlay-visible");
                    existingBackBtn.classList.add("visible");
                }}, 25);
            }});

            // Event listener for Add PO button
            addPoButton.addEventListener('click', async () => {{
                const projectName = poProjectName.textContent;
                const poName = `PO${Date.now()}`;
                
                const result = await window.pywebview.api.create_po(projectName, poName);
                if (result.success) {{
                    const poResult = await window.pywebview.api.get_po_sheets(projectName);
                    if (poResult.success) {{
                        renderPOList(poResult.sheets);
                    }}
                }} else {{
                    console.error("Failed to create PO:", result.message);
                }}
            }});

            // Event listener for Open Excel button
            document.getElementById('openExcelButton').addEventListener('click', async () => {{
                const projectName = poProjectName.textContent;
                const result = await window.pywebview.api.open_excel_file(projectName);
                if (!result.success) {{
                    console.error("Failed to open Excel file:", result.message);
                }}
            }});

            // Window control buttons
            document.getElementById('minimizeButton').addEventListener('click', () => {{
                window.pywebview.api.minimize_window();
            }});

            document.getElementById('maximizeButton').addEventListener('click', () => {{
                window.pywebview.api.toggle_maximize();
            }});

            document.getElementById('closeButton').addEventListener('click', () => {{
                window.pywebview.api.close_window();
            }});
        }});
    </script>
</body>
</html>
"""

def start_app():
    # Expose the API to JavaScript
    window = webview.create_window(
        "Purchase",
        html=html_content,
        js_api=api,
        width=1320,
        height=810,
        resizable=True,
        frameless=True,
        min_size=(400, 300),
        easy_drag=False,  # Disable easy_drag as we're implementing custom dragging
    )
    api.set_window(window)  # Set the window reference
    webview.start(debug=False)

if __name__ == "__main__":
    start_app()